import sys
import os
import time
import subprocess
from datetime import datetime
import tkinter as tk
import openpyxl
from tkinter import colorchooser
import math
from openpyxl.styles import Font
import shutil


class App:

    def __init__(self, root):
        self.root = root
        self.root.title(f'{datetime.now().date()}')
        self.root.minsize(350, 250)
        self.index = 1

        self.attendance = []

        self.label = tk.Label(root, text=f'Attendance for \n {datetime.now().date()}', font=("Arial", 30))
        self.label.pack(pady=30, padx=25)

        self.proceed = tk.Button(text='Real-time', font=("Arial", 20), height=1, width=11, command=self.proceed,
                                 bg='Grey')
        self.proceed.pack(pady=25)

        self.manual = tk.Button(text='Manual', font=("Arial", 20), height=1, width=11, command=self.manual,
                                bg='Pink')
        self.manual.pack()

        self.final = tk.Button(text='Create Final?', font=("Arial", 20), bg='yellow', command=self.bye)
        self.final.pack(pady=25)

        self.workbook = openpyxl.load_workbook('namelist.xlsx')
        self.worksheet = self.workbook.active
        self.empty_column = self.worksheet.max_column + 1

        self.roll = [cell.value for cell in self.worksheet['A']]
        self.names = [cell.value for cell in self.worksheet['B']]

        self.root.bind('<KeyPress>', self.on_key)

    def next_student(self):
        if self.index < len(self.roll):
            student = f'{self.roll[self.index]}, {self.names[self.index]}'
            self.label.config(text=student)
            self.index += 1
        else:
            self.root.destroy()

    def bye(self):
        self.root.destroy()
        app = Total()
        sys.exit()

    def present(self):
        self.attendance.append('Present')
        self.next_student()

    def absent(self):
        self.attendance.append('Absent')
        self.next_student()

    def proceed(self):

        self.proceed.destroy()
        self.manual.destroy()
        self.final.destroy()

        self.present = tk.Button(text='Present', font=("Arial", 20), height=1, width=7, command=self.present,
                                 activebackground="green", bg="green", fg='White')
        self.present.pack(pady=20)

        self.absent = tk.Button(text='Absent', font=("Arial", 20), height=1, width=7, command=self.absent,
                                activebackground="red", bg="red", fg='White')
        self.absent.pack(pady=20)

        self.next_student()

    def append(self):

        new_column = ['', str(datetime.now().date())] + self.attendance

        for iterable in range(1, len(new_column)):
            cell = self.worksheet.cell(row=iterable, column=self.empty_column)

            if iterable == 1:
                cell.font = Font(bold=True)

            cell.value = new_column[iterable]
        if len(new_column) > 2:
            path = f'{get_script_folder()}\\namelist.xlsx'
            self.workbook.save(path)
            new = Total()

    def on_key(self, event):

        if event.keysym == 'Return':
            if self.proceed.winfo_exists():
                self.proceed.invoke()
            else:
                self.present.invoke()
        if event.keysym == 'BackSpace':
            self.absent.invoke()

    def on_enter(self, event):
        if event.keysym == 'Return':
            self.sub.invoke()

    def manual(self):

        self.proceed.destroy()
        self.manual.destroy()
        self.final.destroy()
        self.label.destroy()

        self.root.unbind_all('<KeyPress>')
        self.root.bind('<KeyPress>', self.on_enter)

        self.instr = tk.Label(text='Enter the roll numbers \n of the people who are present', font=('Arial', 24))
        self.instr.pack(pady=20, padx=20)

        self.field = tk.Entry(font=('Times New Roman', 16), width=30)
        self.field.pack(ipady=10, pady=10)

        self.dynamic = tk.Label(text='', font=('Helvetica', 16))
        self.dynamic.pack()

        self.field.bind('<KeyRelease>',self.event_listener)

        self.sub = tk.Button(text='Submit', font=('Arial', 16), fg='white', bg='green', command=self.get_attendance)
        self.sub.pack(pady=20)

    def event_listener(self,event):
        entry_text = self.field.get()
        array = entry_text.split(',')
        display = 'Roll Numbers Entered:'
        for i in array:
            display += ' • '
            display += str(i)
            if array.index(i) % 5 == 0:
                display += '\n'
        self.dynamic.config(text=display,fg='Blue')

    def get_attendance(self):
        text = self.field.get()
        array = text.split(',')

        roll_numbers = list(self.worksheet['A'])
        roll_numbers = roll_numbers[1:len(roll_numbers)]

        for i in roll_numbers:
            cell = self.worksheet.cell(row=int(i.value) + 1, column=self.empty_column)
            cell.value = 'Absent'

        for i in array:
            cell = self.worksheet.cell(row=int(i) + 1, column=self.empty_column)
            cell.value = 'Present'

        cell = self.worksheet.cell(row=1, column=self.empty_column)
        cell.value = f'{datetime.now().date()}'
        cell.font = Font(bold='True')

        self.workbook.save('namelist.xlsx')
        self.root.destroy()


class Total:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title('Generate Namelist?')

        self.workbook = openpyxl.load_workbook('namelist.xlsx')
        self.worksheet = self.workbook.active

        self.max_row = self.worksheet.max_row + 1
        self.max_col = self.worksheet.max_column + 1

        self.attendance_list = []
        for i in range(2, self.max_row):
            self.total = 0
            self.attendance = 0
            for j in range(3, self.max_col):
                self.value = self.worksheet.cell(row=i, column=j).value
                self.total += 1
                if self.value == 'Present':
                    self.attendance += 1
            self.attendance_list.append(self.attendance)

        self.percentages = []
        for i in self.attendance_list:
            percentage = math.ceil(i / self.total * 100)
            self.percentages.append(percentage)

        self.new_column = ['', 'Attendance'] + self.percentages

        self.header = tk.Label(self.root, text='Do you want to generate\n Final Attendance Sheet?', font=('Arial', 32))
        self.header.pack(padx=20, pady=20)

        self.yes = tk.Button(self.root, text='Yes (Generates final.xlsx)', height=2, bg='Blue', fg='White',
                             font=("Arial", 14), command=self.append)
        self.yes.pack(pady=20, padx=20)

        self.no = tk.Button(self.root, text='No (Exits the Program)', height=2, width=19, bg='Red', fg='White',
                            font=("Arial", 14), command=lambda: self.root.destroy())
        self.no.pack(pady=20, padx=20)

        self.root.mainloop()

    def append(self):
        for i in range(1, len(self.new_column)):
            cell = self.worksheet.cell(row=i, column=self.max_col)
            cell.value = self.new_column[i]
            if i == 1:
                cell.font = Font(bold=True)

        path = f'{get_script_folder()}\\final.xlsx'
        self.workbook.save(path)
        self.root.destroy()


class NewWorkbook:

    def __init__(self):
        self.root = tk.Tk()
        self.root.title('Namelist Template Creator')

        self.header = tk.Label(text='You should add a Namelist file', font=('Times New Roman', 20))
        self.header.pack(pady=20, padx=20)

        self.sub = tk.Label(text="Let's create one? Tap the button below", font=('Times New Roman', 12))
        self.sub.pack()

        self.button = tk.Button(text='Create!', font=('Arial', 16), bg='Green', fg='White', command=self.create)
        self.button.pack(pady=20)

        self.root.mainloop()

    def create(self):
        self.workbook = openpyxl.Workbook()
        self.worksheet = self.workbook.active

        cell = self.worksheet.cell(row=1, column=1)
        cell.value = 'ROLL_NO'
        cell.font = Font(bold=True)
        cell = self.worksheet.cell(row=1, column=2)
        cell.font = Font(bold=True)
        cell.value = 'NAME'

        self.worksheet.column_dimensions['A'].width = 10
        self.worksheet.column_dimensions['B'].width = 40

        path = f'{get_script_folder()}\\namelist.xlsx'
        self.workbook.save(path)
        self.root.destroy()


def get_script_folder():
    # path of main .py or .exe when converted with pyinstaller
    if getattr(sys, 'frozen', False):
        script_path = os.path.dirname(sys.executable)
    else:
        script_path = os.path.dirname(
            os.path.abspath(sys.modules['__main__'].__file__)
        )
    return script_path


if __name__ == '__main__':
    script = get_script_folder()

    path = f'{script}\\namelist.xlsx'
    if os.path.exists(path):
        root = tk.Tk()
        app = App(root)
        root.mainloop()
        app.append()
    else:
        app = NewWorkbook()
        time.sleep(3)
        subprocess.Popen(['start', path], shell=True)
        time.sleep(4)
        root = tk.Tk()
        root.title('Info!')
        tk.Label(text='Add Roll Numbers & Names of Students\n to Continue', font=('Times New Roman', 12)).pack(padx=20,
                                                                                                               pady=20)


        def click():
            root.destroy()


        tk.Button(text='OK', font=('Arial', 12), command=click).pack(pady=20)
        root.mainloop()
